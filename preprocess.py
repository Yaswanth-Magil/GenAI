import openpyxl
import json
from collections import defaultdict
import google.generativeai as genai
import os


def aggregate_counts(counts, api_key):
    """Aggregates counts of similar dishes or staff names using Gemini's understanding."""
    aggregated_counts = defaultdict(int)
    processed = set()

    for item1 in counts:
        if item1 in processed:
            continue

        total_count = counts[item1]
        processed.add(item1)

        for item2 in counts:
            if item2 != item1 and item2 not in processed:
                similarity = check_similarity(item1, item2, api_key)

                if similarity >= 0.8:  # Adjust similarity threshold as needed
                    total_count += counts[item2]
                    processed.add(item2)
                    print(f"Aggregating '{item1}' and '{item2}' (Similarity: {similarity:.2f})")

        aggregated_counts[item1] = total_count

    return aggregated_counts


def check_similarity(item1, item2, api_key):
    """Checks the similarity between two items using Gemini."""
    prompt = f"""You are an expert in language semantics.  Determine the semantic similarity between the two phrases provided below, taking into account potential misspellings and variations in phrasing.  Return a score between 0 and 1, where 0 means completely dissimilar and 1 means essentially the same.

Phrase 1: {item1}
Phrase 2: {item2}

Provide only a number (the similarity score) with no other text."""

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash')
        response = model.generate_content(prompt)
        try:
            similarity = float(response.text.strip())
            if 0 <= similarity <= 1:
                return similarity
            else:
                print(f"Warning: Similarity score out of range: {similarity}")
                return 0  # Treat out-of-range scores as dissimilar
        except ValueError:
            print(f"Error: Could not parse similarity score: {response.text}")
            return 0  # Treat unparsable responses as dissimilar
    except Exception as e:
        print(f"Error during similarity check: {e}")
        return 0  # Treat errors as dissimilar


def summarize_reviews(reviews, sentiment_type, api_key):
    """Summarizes positive or negative reviews using Generative AI."""
    if not reviews:
        return "No reviews to summarize.", "N/A"

    prompt = f"""You are an expert in summarizing restaurant reviews. Create a concise summary of the following {sentiment_type} restaurant reviews, highlighting key themes and recurring issues. Format the summary as a list of points, each focusing on a specific aspect like food quality, service, ambiance, etc.  Also, provide a justification for each point in your summary, citing specific excerpts or quotes from the reviews to support your claim.

Here are the {sentiment_type} reviews:
{reviews}"""

    try:
        genai.configure(api_key=api_key)  # Configure Gemini API *within* the function
        model = genai.GenerativeModel('gemini-2.0-flash')
        response = model.generate_content(prompt)
        return response.text.strip(), prompt
    except Exception as e:
        print(f"Error during API call for summarization: {e}")
        return "Error generating summary.", "API Error"


def analyze_competition(my_reviews, competitor_reviews, my_name, competitor_name, api_key):
    """Analyzes reviews to identify areas where each restaurant excels."""

    if not my_reviews or not competitor_reviews:
        return "Insufficient data for comparison.", "N/A", "Insufficient data for comparison.", "N/A"

    prompt_better = f"""You are an expert restaurant reviewer. Compare the following customer reviews of {my_name} with customer reviews of {competitor_name}. Identify specific aspects where {my_name} excels compared to {competitor_name}, such as food quality, service, ambiance, or value. Provide specific examples from the reviews as justification.

{my_name} Reviews:\n{my_reviews}\n\n{competitor_name} Reviews:\n{competitor_reviews}

Focus on highlighting the strengths of {my_name} based on these reviews."""

    prompt_worse = f"""You are an expert restaurant reviewer. Compare the following customer reviews of {my_name} with customer reviews of {competitor_name}. Identify specific aspects where {competitor_name} excels compared to {my_name}, such as food quality, service, ambiance, or value. Provide specific examples from the reviews as justification.

{my_name} Reviews:\n{my_reviews}\n\n{competitor_name} Reviews:\n{competitor_reviews}

Focus on highlighting the strengths of {competitor_name} based on these reviews."""

    try:
        genai.configure(api_key=api_key)  # Configure Gemini API *within* the function
        model = genai.GenerativeModel('gemini-2.0-flash')

        response_better = model.generate_content(prompt_better)
        my_better = response_better.text.strip()

        response_worse = model.generate_content(prompt_worse)
        competitor_better = response_worse.text.strip()

        return my_better, prompt_better, competitor_better, prompt_worse

    except Exception as e:
        print(f"Error during competition analysis: {e}")
        return "Error analyzing competition.", "API Error", "Error analyzing competition.", "API Error"


def process_excel_and_extract_data(input_file_path, output_file_path, api_key):
    """Processes the input Excel file, extracts data, and writes to a new Excel file."""

    competitors = {
        'SPF': 'Chand Palace',
        'Princeton': 'Saravana Bhavan',
        'Parsippany': 'Sangeetha',
        'Chicago': 'Udupi Palace'
    }

    workbook = openpyxl.load_workbook(input_file_path)
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    header = [
        'Outlet', 'Overall Positive Count', 'Overall Negative Count', 'Overall Neutral Count',
        'Dish Positive Count', 'Dish Negative Count', 'Staff Positive Count', 'Staff Negative Count',
        'Category Positive Count', 'Category Negative Count', 'Positive Summary',
        'pos_summary_justification', 'Negative Summary', 'neg_summary_justification',
        'Where_I_do_better', 'Where_I_do_better_justification',
        'Where_competitor_do_better', 'Where_competitor_do_better_justification'
    ]
    output_sheet.append(header)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Processing outlet: {sheet_name}")

        outlet = sheet_name
        overall_positive_count = 0
        overall_negative_count = 0
        overall_neutral_count = 0

        dish_positive_counts = defaultdict(int)
        dish_negative_counts = defaultdict(int)
        staff_positive_counts = defaultdict(int)
        staff_negative_counts = defaultdict(int)
        category_positive_counts = defaultdict(int)
        category_negative_counts = defaultdict(int)

        positive_reviews_for_summary = []
        negative_reviews_for_summary = []

        # Collect competitor reviews *before* the loop
        competitor_positive_reviews = []
        competitor_negative_reviews = []
        found_competitor_sheet = False

        for comp_sheet_name in workbook.sheetnames:
            if competitors.get(sheet_name) == comp_sheet_name:
                competitor_sheet = workbook[comp_sheet_name]
                found_competitor_sheet = True
                print(f"Found competitor sheet: {comp_sheet_name}")

                for row_num, row in enumerate(competitor_sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        review_text = row[5]
                        review_sentiment = row[6]

                        if isinstance(review_sentiment, str):
                            sentiment_lower = review_sentiment.lower()
                            if sentiment_lower == 'positive' and review_text:
                                competitor_positive_reviews.append(review_text)
                            elif sentiment_lower == 'negative' and review_text:
                                competitor_negative_reviews.append(review_text)

                    except Exception as e:
                        print(f"Error processing competitor row {row_num} in sheet {comp_sheet_name}: {e}")
                break  # Stop searching after finding the competitor

        if not found_competitor_sheet:
            print(f"Warning: Competitor sheet '{competitors.get(sheet_name)}' not found.")

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                review_text = row[5]
                review_sentiment = row[6]
                dish_sentiment_str = row[7]
                staff_sentiment_str = row[8]
                category_sentiment_str = row[9]

                if isinstance(review_sentiment, str):
                    sentiment_lower = review_sentiment.lower()
                    if sentiment_lower == 'positive':
                        overall_positive_count += 1
                        if review_text:
                            positive_reviews_for_summary.append(review_text)
                    elif sentiment_lower == 'negative':
                        overall_negative_count += 1
                        if review_text:
                            negative_reviews_for_summary.append(review_text)
                    elif sentiment_lower == 'neutral':
                        overall_neutral_count += 1

                if isinstance(dish_sentiment_str, str) and dish_sentiment_str.strip():
                    try:
                        dish_sentiment = json.loads(dish_sentiment_str)
                        for dish, sentiment in dish_sentiment.items():
                            if sentiment.lower() == 'positive':
                                dish_positive_counts[dish] += 1
                            elif sentiment.lower() == 'negative':
                                dish_negative_counts[dish] += 1
                    except json.JSONDecodeError:
                        print(f"Error parsing dish sentiment in row {row_num}: {dish_sentiment_str}")

                if isinstance(staff_sentiment_str, str) and staff_sentiment_str.strip():
                    try:
                        staff_sentiment = json.loads(staff_sentiment_str)
                        for staff, sentiment in staff_sentiment.items():
                            if sentiment.lower() == 'positive':
                                staff_positive_counts[staff] += 1
                            elif sentiment.lower() == 'negative':
                                staff_negative_counts[staff] += 1
                    except json.JSONDecodeError:
                        print(f"Error parsing staff sentiment in row {row_num}: {staff_sentiment_str}")

                if isinstance(category_sentiment_str, str) and category_sentiment_str.strip():
                    try:
                        category_sentiment = json.loads(category_sentiment_str)
                        for category, sentiment in category_sentiment.items():
                            if sentiment.lower() == 'positive':
                                category_positive_counts[category] += 1
                            elif sentiment.lower() == 'negative':
                                category_negative_counts[category] += 1
                    except json.JSONDecodeError:
                        print(f"Error parsing category sentiment in row {row_num}: {category_sentiment_str}")

            except Exception as e:
                print(f"Error processing row {row_num}: {e}")

        # Aggregate dish counts
        dish_positive_counts_aggregated = aggregate_counts(dish_positive_counts, api_key)
        dish_negative_counts_aggregated = aggregate_counts(dish_negative_counts, api_key)
        staff_positive_counts_aggregated = aggregate_counts(staff_positive_counts, api_key)
        staff_negative_counts_aggregated = aggregate_counts(staff_negative_counts, api_key)
        category_positive_counts_aggregated = aggregate_counts(category_positive_counts, api_key)
        category_negative_counts_aggregated = aggregate_counts(category_negative_counts, api_key)

        # Create positive and negative summaries
        positive_summary, pos_summary_justification = summarize_reviews("\n".join(positive_reviews_for_summary), "positive", api_key)
        negative_summary, neg_summary_justification = summarize_reviews("\n".join(negative_reviews_for_summary), "negative", api_key)

        # Analyze competition
        competitor_name = competitors.get(sheet_name, "Unknown Competitor")
        my_better, my_better_justification, competitor_better, competitor_better_justification = \
            analyze_competition("\n".join(positive_reviews_for_summary + negative_reviews_for_summary),  # All my reviews
                                "\n".join(competitor_positive_reviews + competitor_negative_reviews),  # Competitor reviews
                                sheet_name, competitor_name, api_key)

        # Write the data to the output sheet
        output_row = [
            outlet, overall_positive_count, overall_negative_count, overall_neutral_count,
            json.dumps(dish_positive_counts_aggregated), json.dumps(dish_negative_counts_aggregated),
            json.dumps(staff_positive_counts_aggregated), json.dumps(staff_negative_counts_aggregated),
            json.dumps(category_positive_counts_aggregated), json.dumps(category_negative_counts_aggregated),
            positive_summary, pos_summary_justification, negative_summary, neg_summary_justification,
            my_better, my_better_justification, competitor_better, competitor_better_justification
        ]
        output_sheet.append(output_row)

    output_workbook.save(output_file_path)
    print(f"Processing complete. Data written to {output_file_path}")


# def main():
#     """Main function to execute the data processing."""

#     api_key = "AIzaSyAxk2Wog2ylp7wuQgTGdQCakzJXMoRHzO8"
#     input_file_path = "/Users/yash/Downloads/Today/Splitted/A2b January month.xlsx"
#     output_file_path = "/Users/yash/Downloads/Today/Splitted/output_summary_competitor_analysis.xlsx"

#     process_excel_and_extract_data(input_file_path, output_file_path, api_key)


# if __name__ == "__main__":
#     main()
