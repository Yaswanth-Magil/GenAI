import openpyxl
import time
import os
import json
import traceback
from google.api_core.exceptions import ResourceExhausted
import google.generativeai as genai
from datetime import datetime, date
from dateutil.relativedelta import relativedelta  # Make sure to install: pip install python-dateutil

categories = [
    "Cleanliness", "Menu Variety", "Portion Size", "Staff Friendliness", "Overall Experience",
    "Ambiance", "Speed of Service", "Service", "Value for Money", "Food Quality"
]


def generate_content_from_file(review):
    """Generates sentiment and extracts information from a review using Generative AI model."""
    prompt = f"""You are an expert in analyzing customer reviews for restaurants.  For the following review, please provide the overall sentiment of the review. Also, identify any staff names mentioned along with their sentiment in the review, any dish names mentioned along with their sentiment, and identify sentiment for each matching category from the following list: {', '.join(categories)}.  Provide your response in a JSON format with the following structure:

{{
  "review_sentiment": "positive" or "negative" or "neutral",
  "dish_sentiment": {{"dish_name1": "positive/negative/neutral", "dish_name2": "positive/negative/neutral"}} or {{}},
  "staff_sentiment": {{"staff_name1": "positive/negative/neutral", "staff_name2": "positive/negative/neutral"}} or {{}},
  "category_sentiment": {{"category1": "positive/negative/neutral", "category2": "positive/negative/neutral"}} or {{}}
}}

If a field cannot be determined, set its value to an empty dictionary (for dish_sentiment, staff_sentiment, category_sentiment) or neutral for review_sentiment. Make sure the keys are always enclosed in double quotes.

Here is the review: {review}"""
    max_retries = 5
    for attempt in range(max_retries):
        try:
            response = genai.GenerativeModel('gemini-2.0-flash').generate_content(prompt)
            return response.text.strip()
        except ResourceExhausted as e:
            if attempt < max_retries - 1:
                sleep_time = 9 ** attempt  # Exponential backoff
                print(f"Quota exceeded. Retrying in {sleep_time} seconds...")
                time.sleep(sleep_time)
            else:
                raise e
        except Exception as e:
            print(f"Error during API call: {e}")
            return None  # or raise, depending on your desired behavior


def get_column_index(sheet, column_name):
    """Finds the index of the specified column name."""
    for cell in sheet[1]:
        if cell.value and cell.value.strip().lower() == column_name.lower():
            return cell.column
    return None


def process_reviews(xlsx_file_path):
    """Processes reviews from all sheets in an Excel file and adds sentiment and extractions."""
    workbook = openpyxl.load_workbook(xlsx_file_path)

    for sheet in workbook.worksheets:
        sheet_name = sheet.title
        print(f"Processing sheet: {sheet_name}")

        # Check if columns already exist
        review_sentiment_column_index = get_column_index(sheet, 'Review Sentiment')
        dish_sentiment_column_index = get_column_index(sheet, 'Dish Sentiment')
        staff_sentiment_column_index = get_column_index(sheet, 'Staff Sentiment')
        category_sentiment_column_index = get_column_index(sheet, 'Category Sentiment')
        reviews_column_index = get_column_index(sheet, 'Reviews')

        if not reviews_column_index:
            print(f"Error: 'Reviews' column not found in sheet {sheet_name}. Skipping...")
            continue

        # Add columns if they don't exist
        next_available_column = sheet.max_column + 1

        if not review_sentiment_column_index:
            sheet.cell(row=1, column=next_available_column, value='Review Sentiment')
            review_sentiment_column_index = next_available_column
            next_available_column += 1

        if not dish_sentiment_column_index:
            sheet.cell(row=1, column=next_available_column, value='Dish Sentiment')
            dish_sentiment_column_index = next_available_column
            next_available_column += 1

        if not staff_sentiment_column_index:
            sheet.cell(row=1, column=next_available_column, value='Staff Sentiment')
            staff_sentiment_column_index = next_available_column
            next_available_column += 1

        if not category_sentiment_column_index:
            sheet.cell(row=1, column=next_available_column, value='Category Sentiment')
            category_sentiment_column_index = next_available_column
            next_available_column += 1


        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            review = row[reviews_column_index - 1] if len(row) >= reviews_column_index else None

            if review:
                try:
                    api_response = generate_content_from_file(review)

                    if api_response:
                        print(f"API Response: {api_response}")  # Add this line for debugging

                        # Remove the extra characters before and after the JSON
                        api_response = api_response.replace("```json", "").replace("```", "").strip()

                        try:
                            #api_response = api_response.encode('utf-8').decode('utf-8') #Try to fix decoding errors
                            data = json.loads(api_response)
                            review_sentiment = data.get('review_sentiment', 'Unknown')
                            dish_sentiment = data.get('dish_sentiment', {})
                            staff_sentiment = data.get('staff_sentiment', {})
                            category_sentiment = data.get('category_sentiment', {})

                            sheet.cell(row=row_num, column=review_sentiment_column_index, value=review_sentiment)
                            sheet.cell(row=row_num, column=dish_sentiment_column_index, value=json.dumps(dish_sentiment))
                            sheet.cell(row=row_num, column=staff_sentiment_column_index, value=json.dumps(staff_sentiment))
                            sheet.cell(row=row_num, column=category_sentiment_column_index, value=json.dumps(category_sentiment))

                            print(f"Review: {review}\nReview Sentiment: {review_sentiment}\nDish Sentiment: {dish_sentiment}\nStaff Sentiment: {staff_sentiment}\nCategory Sentiment: {category_sentiment}\n")

                        except json.JSONDecodeError as e:
                            print(f"Error decoding JSON response in sheet {sheet_name} row {row_num}: {e}\nResponse was: {api_response}")
                            traceback.print_exc() #Print the traceback
                            with open("json_error_log.txt", "a") as f: #Log response to a file
                                f.write(f"Sheet: {sheet_name}, Row: {row_num}\n")
                                f.write(f"Response: {api_response}\n")
                                f.write(traceback.format_exc() + "\n")

                            sheet.cell(row=row_num, column=review_sentiment_column_index, value="JSON Error")
                            sheet.cell(row=row_num, column=dish_sentiment_column_index, value="JSON Error")
                            sheet.cell(row=row_num, column=staff_sentiment_column_index, value="JSON Error")
                            sheet.cell(row=row_num, column=category_sentiment_column_index, value="JSON Error")
                        except UnicodeDecodeError as e:
                            print(f"UnicodeDecodeError: {e}")
                            # Handle the encoding error appropriately (e.g., try a different encoding)
                            sheet.cell(row=row_num, column=review_sentiment_column_index, value="Encoding Error")
                            sheet.cell(row=row_num, column=dish_sentiment_column_index, value="Encoding Error")
                            sheet.cell(row=row_num, column=staff_sentiment_column_index, value="Encoding Error")
                            sheet.cell(row=row_num, column=category_sentiment_column_index, value="Encoding Error")


                    else:
                        print(f"No response from API for review in sheet {sheet_name} row {row_num}")
                        sheet.cell(row=row_num, column=review_sentiment_column_index, value="API Error")
                        sheet.cell(row=row_num, column=dish_sentiment_column_index, value="API Error")
                        sheet.cell(row=row_num, column=staff_sentiment_column_index, value="API Error")
                        sheet.cell(row=row_num, column=category_sentiment_column_index, value="API Error")


                except Exception as e:
                    print(f"Error processing review in sheet {sheet_name} row {row_num}: {e}")
                    sheet.cell(row=row_num, column=review_sentiment_column_index, value="Error")
                    sheet.cell(row=row_num, column=dish_sentiment_column_index, value="Error")
                    sheet.cell(row=row_num, column=staff_sentiment_column_index, value="Error")
                    sheet.cell(row=row_num, column=category_sentiment_column_index, value="Error")

            else:
                print("No review text found. Skipping...\n")

    workbook.save(xlsx_file_path)
    print(f"Sentiment analysis and extraction completed. Updated file: {xlsx_file_path}")