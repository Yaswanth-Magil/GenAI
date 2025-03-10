import openpyxl
import json
from collections import defaultdict
import google.generativeai as genai
import os
import ReviewAnalysis2
from collections import OrderedDict
from datetime import date
from dateutil.relativedelta import relativedelta


def aggregate_counts(counts, api_key):
    """Aggregates counts of similar dishes or staff names using Gemini's understanding."""
    aggregated_counts = defaultdict(int)
    processed = set()

    for item1 in counts:
        if item1 in processed:
            continue

        total_count = counts[item1]
        processed.add(item1)

        for item2 in counts:
            if item2 != item1 and item2 not in processed:
                similarity = check_similarity(item1, item2, api_key)

                if similarity >= 0.8:  # Adjust similarity threshold as needed
                    total_count += counts[item2]
                    processed.add(item2)
                    print(f"Aggregating '{item1}' and '{item2}' (Similarity: {similarity:.2f})")

        aggregated_counts[item1] = total_count

    return aggregated_counts


def check_similarity(item1, item2, api_key):
    """Checks the similarity between two items using Gemini."""
    prompt = f"""You are an expert in language semantics.  Determine the semantic similarity between the two phrases provided below, taking into account potential misspellings and variations in phrasing.  Return a score between 0 and 1, where 0 means completely dissimilar and 1 means essentially the same.

Phrase 1: {item1}
Phrase 2: {item2}

Provide only a number (the similarity score) with no other text."""

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash')
        response = model.generate_content(prompt)
        try:
            similarity = float(response.text.strip())
            if 0 <= similarity <= 1:
                return similarity
            else:
                print(f"Warning: Similarity score out of range: {similarity}")
                return 0  # Treat out-of-range scores as dissimilar
        except ValueError:
            print(f"Error: Could not parse similarity score: {response.text}")
            return 0  # Treat unparsable responses as dissimilar
    except Exception as e:
        print(f"Error during similarity check: {e}")
        return 0  # Treat errors as dissimilar


def summarize_reviews(reviews, sentiment_type, api_key):
    """Summarizes positive or negative reviews using Generative AI."""
    if not reviews:
        return "No reviews to summarize.", "N/A"

    prompt = f"""You are an expert in summarizing restaurant reviews. Create a concise summary of the following {sentiment_type} restaurant reviews, highlighting key themes and recurring issues. Format the summary as a list of points, each focusing on a specific aspect like food quality, service, ambiance, etc.  Also, provide a justification for each point in your summary, citing specific excerpts or quotes from the reviews to support your claim.

Here are the {sentiment_type} reviews:
{reviews}"""

    try:
        genai.configure(api_key=api_key)  # Configure Gemini API *within* the function
        model = genai.GenerativeModel('gemini-2.0-flash')
        response = model.generate_content(prompt)
        return response.text.strip(), prompt
    except Exception as e:
        print(f"Error during API call for summarization: {e}")
        return "Error generating summary.", "API Error"


def analyze_competition(my_reviews, competitor_reviews, my_name, competitor_name, api_key):
    """Analyzes reviews to identify areas where each restaurant excels."""

    if not my_reviews or not competitor_reviews:
        return "Insufficient data for comparison.", "N/A", "Insufficient data for comparison.", "N/A"

    prompt_better = f"""You are an expert restaurant reviewer. Compare the following customer reviews of {my_name} with customer reviews of {competitor_name}. Identify specific aspects where {my_name} excels compared to {competitor_name}, such as food quality, service, ambiance, or value. Provide specific examples from the reviews as justification.

{my_name} Reviews:\n{my_reviews}\n\n{competitor_name} Reviews:\n{competitor_reviews}

Focus on highlighting the strengths of {my_name} based on these reviews."""

    prompt_worse = f"""You are an expert restaurant reviewer. Compare the following customer reviews of {my_name} with customer reviews of {competitor_name}. Identify specific aspects where {competitor_name} excels compared to {my_name}, such as food quality, service, ambiance, or value. Provide specific examples from the reviews as justification.

{my_name} Reviews:\n{my_reviews}\n\n{competitor_name} Reviews:\n{competitor_reviews}

Focus on highlighting the strengths of {competitor_name} based on these reviews."""

    try:
        genai.configure(api_key=api_key)  # Configure Gemini API *within* the function
        model = genai.GenerativeModel('gemini-2.0-flash')

        response_better = model.generate_content(prompt_better)
        my_better = response_better.text.strip()

        response_worse = model.generate_content(prompt_worse)
        competitor_better = response_worse.text.strip()

        return my_better, prompt_better, competitor_better, prompt_worse

    except Exception as e:
        print(f"Error during competition analysis: {e}")
        return "Error analyzing competition.", "API Error", "Error analyzing competition.", "API Error"


def analyze_trend_shift(previous_month_reviews, current_month_reviews, outlet, trend_type, api_key):
    """Analyzes reviews to identify reasons for trend shifts."""

    if not previous_month_reviews or not current_month_reviews:
        return "Insufficient data to determine trend shifts."

    prompt = f"""You are an expert in analyzing restaurant review trends. Identify the top 3 reasons why customer reviews for {outlet} shifted from {trend_type.split('_')[1]} in the previous month to {trend_type.split('_')[3]} in the current month. Provide specific examples from the reviews as justification.

Previous Month Reviews:\n{previous_month_reviews}\n\nCurrent Month Reviews:\n{current_month_reviews}

Focus on the most significant changes in customer sentiment."""

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash')
        response = model.generate_content(prompt)
        return response.text.strip()

    except Exception as e:
        print(f"Error during trend shift analysis: {e}")
        return "Error analyzing trend shift."


def process_excel_and_extract_data(input_file_path, output_file_path, api_key):
    """Processes the input Excel file, extracts data, and writes to a new Excel file."""

    competitors = {
        # 'SPF': 'Chand Palace',
        'Princeton': 'Saravana Bhavan',
        # 'Parsippany': 'Sangeetha',
        # 'Chicago': 'Udupi Palace'
    }

    workbook = openpyxl.load_workbook(input_file_path)
    # Assuming everything is in the first sheet
    sheet = workbook.active

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    header = [
        'Outlet', 'Overall Positive Count', 'Overall Negative Count', 'Overall Neutral Count',
        'Dish Positive Count', 'Dish Negative Count', 'Staff Positive Count', 'Staff Negative Count',
        'Category Positive Count', 'Category Negative Count', 'Positive Summary',
        'pos_summary_justification', 'Negative Summary', 'neg_summary_justification',
        'Where_I_do_better', 'Where_I_do_better_justification',
        'Where_competitor_do_better', 'Where_competitor_do_better_justification',
        'Trend_Pos_To_Neg', 'Trend_Neg_To_Pos'  # Added trend columns
    ]
    output_sheet.append(header)

    # ***FIX: Getting column indices for main sheet and outlet***
    outlet_col_index = ReviewAnalysis2.get_column_index(sheet, 'Outlet')
    review_text_col_index = ReviewAnalysis2.get_column_index(sheet, 'Reviews')
    review_sentiment_col_index = ReviewAnalysis2.get_column_index(sheet, 'Review Sentiment')
    dish_sentiment_col_index = ReviewAnalysis2.get_column_index(sheet, 'Dish Sentiment')
    staff_sentiment_col_index = ReviewAnalysis2.get_column_index(sheet, 'Staff Sentiment')
    category_sentiment_col_index = ReviewAnalysis2.get_column_index(sheet, 'Category Sentiment')

    if not outlet_col_index or not review_text_col_index or not review_sentiment_col_index or not dish_sentiment_col_index or not staff_sentiment_col_index or not category_sentiment_col_index:
        print(f"Error: One or more required columns ('Outlet', 'Reviews', 'Review Sentiment', 'Dish Sentiment', 'Staff Sentiment', 'Category Sentiment') not found.  Skipping processing.")
        return

    # Store reviews by outlet and month (using OrderedDict to maintain order)
    outlet_reviews = defaultdict(lambda: OrderedDict())
    
    #Collect all previous and following months data for current previous month.
    
    for outlet in competitors.keys():  # Changed to iterate over known outlets
        #Get first previous month
        allReviews = defaultdict(lambda: {'positive': [], 'negative': []})

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                current_outlet = row[outlet_col_index - 1] if outlet_col_index and len(row) >= outlet_col_index else None
                review_text = row[review_text_col_index - 1] if review_text_col_index and len(row) >= review_text_col_index else None
                review_sentiment = row[review_sentiment_col_index - 1] if review_sentiment_col_index and len(row) >= review_sentiment_col_index else None

                if not current_outlet or not review_text or not review_sentiment:
                    continue  # Skip rows with missing data

                #Only process for the selected outlet 
                if current_outlet.strip() != outlet:
                    continue
                    
                if isinstance(review_sentiment, str):
                    sentiment_lower = review_sentiment.lower()
                    if sentiment_lower == 'positive':
                        allReviews[current_outlet]['positive'].append(review_text)
                    elif sentiment_lower == 'negative':
                        allReviews[current_outlet]['negative'].append(review_text)

            except Exception as e:
                print(f"Error collecting review data in row {row_num}: {e}")
                continue
        
        print(f"Processing outlet: {outlet}")
        
        #Checking if previous month exists before doing any call to Gemini
        if not allReviews or not allReviews[outlet]['positive'] or not allReviews[outlet]['negative']:
            print(f"Skipping outlet {outlet}: No reviews found.")
            output_row = [outlet] + ["N/A"] * 19 #Output N/A if outlet had no review
            output_sheet.append(output_row)
            continue #Skip if there is not previous review
            
        #We have data, so Lets create variables
        overall_positive_count = len(allReviews[outlet]['positive'])
        overall_negative_count = len(allReviews[outlet]['negative'])
        overall_neutral_count = 0 #As the neutral was unused I put it as 0.
            
        # Aggregate dish counts, which for the purpose is empty
        dish_positive_counts = defaultdict(int)
        dish_negative_counts = defaultdict(int)
        staff_positive_counts = defaultdict(int)
        staff_negative_counts = defaultdict(int)
        category_positive_counts = defaultdict(int)
        category_negative_counts = defaultdict(int)
        
        dish_positive_counts_aggregated = aggregate_counts(dish_positive_counts, api_key)
        dish_negative_counts_aggregated = aggregate_counts(dish_negative_counts, api_key)
        staff_positive_counts_aggregated = aggregate_counts(staff_positive_counts, api_key)
        staff_negative_counts_aggregated = aggregate_counts(staff_negative_counts, api_key)
        category_positive_counts_aggregated = aggregate_counts(category_positive_counts, api_key)
        category_negative_counts_aggregated = aggregate_counts(category_negative_counts, api_key)
        
        #Now get summaries
        positive_reviews_for_summary = allReviews[outlet]['positive']
        negative_reviews_for_summary = allReviews[outlet]['negative']
        
        # Create positive and negative summaries (using all reviews)
        positive_summary, pos_summary_justification = summarize_reviews("\n".join(positive_reviews_for_summary), "positive", api_key)
        negative_summary, neg_summary_justification = summarize_reviews("\n".join(negative_reviews_for_summary), "negative", api_key)

        # Analyze competition (skipping for now)
        my_better, my_better_justification, competitor_better, competitor_better_justification = "N/A", "N/A", "N/A", "N/A"
        
        #Start to create the final output
        
        #Get trends.
        trend_pos_to_neg = ""
        trend_neg_to_pos = ""
        
        #The trend shift is not correct.
        
        # Write the data to the output sheet
        output_row = [
            outlet, overall_positive_count, overall_negative_count, overall_neutral_count,
            json.dumps(dish_positive_counts_aggregated), json.dumps(dish_negative_counts_aggregated),
            json.dumps(staff_positive_counts_aggregated), json.dumps(staff_negative_counts_aggregated),
            json.dumps(category_positive_counts_aggregated), json.dumps(category_negative_counts_aggregated),
            positive_summary, pos_summary_justification, negative_summary, neg_summary_justification,
            my_better, my_better_justification, competitor_better, competitor_better_justification,
            trend_pos_to_neg, trend_neg_to_pos  # Added trend data
        ]
        output_sheet.append(output_row)
    
    output_workbook.save(output_file_path)
    print(f"Processing complete. Data written to {output_file_path}")